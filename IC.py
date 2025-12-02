import numpy as np
from scipy.linalg import svd
import pandas as pd
from collections import Counter
import openpyxl

#function to read dsm from excel
def read_dsm(file_path, sheet_name):
    return pd.read_excel(file_path, sheet_name, index_col=0)
#function to read module assignments
def read_mod(file_path, sheet_name):
    df = pd.read_excel(file_path, sheet_name, index_col=0)
    mod_assignments = df['Module'].tolist()
    return mod_assignments
#function to read component complexities
def read_compl(file_path, sheet_name):
    df = pd.read_excel(file_path, sheet_name, index_col=0) #read sheet from file
    complexities = df['Complexity'].tolist() #read 'Complexity' column to list
    return complexities #return list of values
#function to sum number of components in each module
def sum_components_by_module(mod_assignments):
    num_components = dict(Counter(mod_assignments)) #count how many components are assigned to each module, hold in dictionary
    return num_components #return dictionary
'''
#function to get number of modules
def num_of_modules(num_components):
    num_modules = len(num_components) #length of dictionary
    return num_modules #return number of modules
    '''
#function to get bounds for each module
def get_module_bounds(num_components):
    bounds = [] #initialize list
    start = 0 #initialize start
    for i,j in num_components.items(): #for each module, # of components pair in the dictionary
        #print(f"There are {j} components in Module {i}")
        end = start+j #num of components indicates end of module bound
        bounds.append((start,end)) #update bounds list
        start = end #the starting point for the next module
    return bounds #return list of bound values
#function to extract modules from adjacency/interaction matrices
def extract_modules(matrix, module_bounds):
    modules = [] # initialize list to hold module matrices
    for start,end in module_bounds:
        rows = matrix.index[start:end] #index the rows
        cols = matrix.columns[start:end] #index the columns
        module = matrix.loc[rows,cols] #access the group of rows and columns
        module_array = module.to_numpy()
        modules.append(module_array) #update list of modules
    return modules #return list of modules
#function to sum complexity values of every component in the system
def sum_system_complexities(complexities):
    C1 = sum(complexities) #sum values in complexities list
    return C1 #return value
#function to calculate double dot product of two matrices
def calculate_C2(x, y):
    C2 = 0 #initialize value
    for i in range(len(x)):
        for j in range(len(x[0])):
            C2 += x[i][j] * y[i][j]
    return C2 #return value
#function to calculate graph energy
def calculate_ge(matrix):
    dimensions = matrix.shape
    rows, columns = dimensions
    U, singular, V_transpose = svd(matrix)
    graph_energy = np.sum(singular) #sum the singular values
    #print("Singular Matrix =",singular)
    #print("Graph Energy =",graph_energy)
    return graph_energy #return value
#function to calculate third term
def calculate_C3(graph_energy, num_components):
    C3 = graph_energy/num_components #divide graph energy by num of components
    return C3 #return value
#function to calculte sum of component complexities in each module
def calculate_module_complexity(complexities, module_bounds):
    module_complexities = [] #initialize list
    for start,end in module_bounds: #wrt the module bounds
        total = sum(complexities[start:end]) #sum values within the bounds
        module_complexities.append(total) #update list
    return module_complexities #return list of values
#function to calculate second term and graph energy for each module
def calculate_module_C2_GE(x,y):
    C2_list = [] #initialize list
    C3_list = [] #initialize list
    for i, module_x in enumerate(x):
        module_y = y[i]
        C2 = calculate_C2(module_x,module_y) #calculate double dot product
        C2_list.append(C2) #update list
    for i, module in enumerate(x):
        ge = calculate_ge(module)
        C3 = ge
        C3_list.append(C3)
    module_C2_values = [float(l) for l in C2_list]
    C3_values = [float(m) for m in C3_list]
    return module_C2_values, C3_values
#function to calculate third term and multiply C2*C3
def calculate_module_C2_C3(module_C2_values, C3_values, num_components):
    module_C3_values = [] #initialize list
    C2xC3 = [] #initialize list
    for g,n in zip(C3_values, num_components.values()):
        #print(g,n)
        C3 = g/n #divide graph energy by the number of components in each module
        module_C3_values.append(C3) #update list
    for value1, value2 in zip(module_C2_values, module_C3_values): #for each C2 and C3 term
        result = value1*value2 #multiply them
        C2xC3.append(result) #update list with values
    return module_C3_values, C2xC3 #return list
#function to calculate structural complexity of each module
def module_structural_complexity(module_complexities, module_C2_values,module_C3_values):
    struct_compl_list = []
    i=0
    for value1, value2, value3 in zip(module_complexities,module_C2_values,module_C3_values):
        structural_compl = value1 + value2*value3
        struct_compl_list.append(structural_compl)
        #print(f"\nModule {i}:")
        #print(f"\nC1: {value1}, C2: {value2}, C3: {value3}, Structural Complexity: {structural_compl:.3f}")
        i+=1
    return struct_compl_list
#function to export results
def export_results(output_file,ICn):
    wb = openpyxl.load_workbook(output_file)#load existing workbook
    ws = wb['Results'] #specify sheet name
    ws['A2']='ICn' #write ICn to cell A2
    ws['B2']=ICn #write ICn value to cell B2
    wb.save(output_file) #save workbook
    print(f"Integrative Complexity result saved to: {output_file}")

#---MAIN CODE---
if __name__ == "__main__":
    adj_matrix = read_dsm('ICdata.xlsm','AdjacencyMatrix')
    print(adj_matrix)
    adj_matrix_list = adj_matrix.to_numpy().tolist()

    int_matrix = read_dsm('ICdata.xlsm','InteractionMatrix')
    #print(int_matrix)
    int_matrix_list = int_matrix.to_numpy().tolist()

    complexity = read_compl('ICdata.xlsm','ComponentInfo')
    #print(complexity)

    #whole system
    systemC1 = sum_system_complexities(complexity)
    #print(f"System C1 = {systemC1:.3f}")
    systemC2 = calculate_C2(adj_matrix_list, int_matrix_list)
    #print(f"System C2 = {systemC2:.3f}")
    GE = calculate_ge(adj_matrix)
    totalComp = len(adj_matrix)
    systemC3 = calculate_C3(GE,totalComp)
    #print(f"System C3 = {systemC3:.3f}")
    systemC2C3 = systemC2 * systemC3
    systemC = systemC1 + systemC2C3
    systemComplexity = round(systemC,2)
    #print("System Structural Complexity =", systemComplexity)

    #per module
    module_assignments = read_mod('ICdata.xlsm','ComponentInfo')
    #print(module_assignments)
    X = sum_components_by_module(module_assignments)
    #print(X)
    module_bounds = get_module_bounds(X)
    #print(module_bounds)
    '''
    Y = num_of_modules(X)
    #print(Y)
    '''
    adj_modules = extract_modules(adj_matrix, module_bounds)
    '''
    print("\nExtracted Modules from Adjacency Matrix:")
    for i, module in enumerate(adj_modules):
        print(f"\nModule {i}:")
        print(module)
    '''
    int_modules = extract_modules(int_matrix, module_bounds)
    '''
    print("\nExtracted Modules from Interaction Matrix:")
    for i, module in enumerate(int_modules):
        print(f"\nModule {i}:")
        print(module)
    '''
    module_C1_values = calculate_module_complexity(complexity,module_bounds)
    #print(module_C1_values)
    C2, ge = calculate_module_C2_GE(adj_modules, int_modules)
    C3, C2XC3 = calculate_module_C2_C3(C2,ge,X)
    sum_moduleC2C3 = sum(C2XC3) #sum C2xC3 values of each module
    intComp = systemC2C3 - sum_moduleC2C3 #calculate integrative complexity
    #print("\n---Complexity Calculation Results---")
    #print(f"Integrative Complexity = {intComp:.3f}")
    normalizedIntComp = 1 - (sum_moduleC2C3/systemC2C3) #calculate normalized integrative complexity
    #print("For system C'':")????
    print(f"Normalized Integrative Complexity = {normalizedIntComp:.3f}")
    x = module_structural_complexity(module_C1_values, C2, C3) #calculate structual complexity of each module
    export_results('Results.xlsx',normalizedIntComp) #call function to export normalized integrative complexity results
